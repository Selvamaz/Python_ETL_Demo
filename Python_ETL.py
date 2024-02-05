import pandas as pd
import psycopg2
import os
from openpyxl import load_workbook

All_Row_Data = []

try:
    connected_excel = 'C://Users//selva/OneDrive//Desktop//supermarket_sales.xlsx'
    connected_wb = load_workbook(connected_excel)
    connected_Ws = connected_wb['supermarket_sales']
    All_Rows = list(connected_Ws.rows)
    cell = [column.value for column in All_Rows[0]]
    print("Excel has been read")
except:
    print("Error Reading Excel file !!!")

for row in connected_Ws.iter_rows(): 
    All_Row_Data.append(row)

try:
    connected_db = psycopg2.connect(database ="ETLDemo", host = "localhost", user = "postgres", password = "5388", port = "5432")
    connected_db.autocommit = True
    print("Succefully connected to the database !!!")
except:
    print("Error in connecting !!!")

cursordb = connected_db.cursor()
new_table_name = "people12"
table_finder = False

cursordb.execute("""SELECT table_name FROM information_schema.tables
       WHERE table_schema = 'public'""")
for table in cursordb.fetchall():
    table_name = table[0]
    if(table_name == new_table_name):
        table_finder = True
        create_table = table_name
        print("Table already exists with the same name !!")
    
if table_finder == False:
    create_table = f"""CREATE TABLE {new_table_name}(
                        {", ". join([f'"{name}" VARCHAR(200)' for name in cell])})"""
    try:
        cursordb.execute(create_table)
        print("Successfull Table Creation")
    except:
        print("Unsuccessfull Table Creation")
    try:
        for row in All_Rows[1:]:
            insert_query = f"""INSERT INTO {new_table_name} values('{str(row[0].value)}','{str(row[1].value)}','{str(row[2].value)}','{str(row[3].value)}','{str(row[4].value)}','{str(row[5].value)}','{str(row[6].value)}','{str(row[7].value)}','{str(row[8].value)}','{str(row[9].value)}','{str(row[10].value)}','{str(row[11].value)}','{str(row[12].value)}','{str(row[13].value)}','{str(row[14].value)}','{str(row[15].value)}','{str(row[16].value)}')"""
            cursordb.execute(insert_query)
        print("Successfull Table Entry")
    except:
        print("Unsuccessfull Table Entry")  

    
else:
    print("Please try other name !!!")

  